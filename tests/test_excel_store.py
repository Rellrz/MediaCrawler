# -*- coding: utf-8 -*-
# Copyright (c) 2025 relakkes@gmail.com
#
# This file is part of MediaCrawler project.
# Repository: https://github.com/NanmiCoder/MediaCrawler/blob/main/tests/test_excel_store.py
# GitHub: https://github.com/NanmiCoder
# Licensed under NON-COMMERCIAL LEARNING LICENSE 1.1
#
# 声明：本代码仅供学习和研究目的使用。使用者应遵守以下原则：
# 1. 不得用于任何商业用途。
# 2. 使用时应遵守目标平台的使用条款和robots.txt规则。
# 3. 不得进行大规模爬取或对平台造成运营干扰。
# 4. 应合理控制请求频率，避免给目标平台带来不必要的负担。
# 5. 不得用于任何非法或不当的用途。
#
# 详细许可条款请参阅项目根目录下的LICENSE文件。
# 使用本代码即表示您同意遵守上述原则和LICENSE中的所有条款。

"""
Unit tests for Excel export functionality
"""

import pytest
import asyncio
import os
from pathlib import Path
import tempfile
import shutil

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from store.excel_store_base import ExcelStoreBase


@pytest.mark.skipif(not EXCEL_AVAILABLE, reason="openpyxl not installed")
class TestExcelStoreBase:
    """Test cases for ExcelStoreBase"""

    @pytest.fixture(autouse=True)
    def clear_singleton_state(self):
        """Clear singleton state before and after each test"""
        ExcelStoreBase._instances.clear()
        yield
        ExcelStoreBase._instances.clear()

    @pytest.fixture
    def temp_dir(self):
        """Create temporary directory for test files"""
        temp_path = tempfile.mkdtemp()
        yield temp_path
        # Cleanup
        shutil.rmtree(temp_path, ignore_errors=True)

    @pytest.fixture
    def excel_store(self, temp_dir, monkeypatch):
        """Create ExcelStoreBase instance for testing"""
        # Monkey patch data directory
        monkeypatch.chdir(temp_dir)
        store = ExcelStoreBase(platform="test", crawler_type="search")
        yield store
        # Cleanup is handled by temp_dir fixture

    def test_initialization(self, excel_store):
        """Test Excel store initialization"""
        assert excel_store.platform == "test"
        assert excel_store.crawler_type == "search"
        assert excel_store.workbooks == {}
        assert set(excel_store.filenames) == {"content", "comment"}
        assert excel_store.get_filename("contents").name.startswith(
            "test_search_content_"
        )
        assert excel_store.get_filename("comments").name.startswith(
            "test_search_comment_"
        )

    @pytest.mark.asyncio
    async def test_store_content(self, excel_store):
        """Test storing content data"""
        content_item = {
            "note_id": "test123",
            "title": "Test Title",
            "desc": "Test Description",
            "user_id": "user456",
            "nickname": "TestUser",
            "liked_count": 100,
            "comment_count": 50
        }

        await excel_store.store_content(content_item)

        # Verify data was written
        assert excel_store.sheets["Contents"].max_row == 2
        assert "Contents" in excel_store.headers_written

    @pytest.mark.asyncio
    async def test_store_comment(self, excel_store):
        """Test storing comment data"""
        comment_item = {
            "comment_id": "comment123",
            "note_id": "note456",
            "content": "Great post!",
            "user_id": "user789",
            "nickname": "Commenter",
            "like_count": 10
        }

        await excel_store.store_comment(comment_item)

        # Verify data was written
        assert excel_store.sheets["Comments"].max_row == 2
        assert "Comments" in excel_store.headers_written

        filename = excel_store.get_filename("comment")
        assert filename.exists()
        workbook = openpyxl.load_workbook(filename)
        assert workbook["Comments"].cell(row=2, column=1).value == "comment123"
        workbook.close()

    @pytest.mark.asyncio
    async def test_store_creator(self, excel_store):
        """Test storing creator data"""
        creator_item = {
            "user_id": "creator123",
            "nickname": "Creator Name",
            "fans": 10000,
            "follows": 500,
            "interaction": 50000
        }

        await excel_store.store_creator(creator_item)

        # Verify data was written
        assert excel_store.sheets["Creators"].max_row == 2
        assert "Creators" in excel_store.headers_written

    @pytest.mark.asyncio
    async def test_multiple_items(self, excel_store):
        """Test storing multiple items"""
        # Store multiple content items
        for i in range(5):
            await excel_store.store_content({
                "note_id": f"note{i}",
                "title": f"Title {i}",
                "liked_count": i * 10
            })

        # Verify all items were stored
        assert excel_store.sheets["Contents"].max_row == 6

    def test_flush(self, excel_store):
        """Test flushing data to file"""
        # Add some test data
        asyncio.run(excel_store.store_content({
            "note_id": "test",
            "title": "Test"
        }))

        # Flush to file
        excel_store.flush()

        # Verify file was created
        filename = excel_store.get_filename("content")
        assert filename.exists()

        # Verify file can be opened
        wb = openpyxl.load_workbook(filename)
        assert "Contents" in wb.sheetnames
        wb.close()

    def test_header_formatting(self, excel_store):
        """Test header row formatting"""
        asyncio.run(excel_store.store_content({"note_id": "test", "title": "Test"}))

        # Check header formatting
        header_cell = excel_store.sheets["Contents"].cell(row=1, column=1)
        assert header_cell.font.bold is True
        # RGB color may have different prefix (00 or FF), check the actual color part
        assert header_cell.fill.start_color.rgb[-6:] == "366092"

    def test_empty_sheets_removed(self, excel_store):
        """Test that empty sheets are removed on flush"""
        # Only add content, leave comments and creators empty
        asyncio.run(excel_store.store_content({"note_id": "test"}))

        excel_store.flush()

        # Reload workbook
        wb = openpyxl.load_workbook(excel_store.get_filename("content"))

        # Only Contents sheet should exist
        assert "Contents" in wb.sheetnames
        assert "Comments" not in wb.sheetnames
        assert "Creators" not in wb.sheetnames
        wb.close()

    @pytest.mark.asyncio
    async def test_content_and_comments_use_separate_files(self, excel_store):
        await excel_store.store_content({"note_id": "note1"})
        await excel_store.store_comment({"comment_id": "comment1"})
        excel_store.flush()

        content_filename = excel_store.get_filename("content")
        comment_filename = excel_store.get_filename("comment")
        assert content_filename != comment_filename
        assert content_filename.exists()
        assert comment_filename.exists()

        content_workbook = openpyxl.load_workbook(content_filename)
        comment_workbook = openpyxl.load_workbook(comment_filename)
        assert content_workbook.sheetnames == ["Contents"]
        assert comment_workbook.sheetnames == ["Comments"]
        content_workbook.close()
        comment_workbook.close()


@pytest.mark.skipif(not EXCEL_AVAILABLE, reason="openpyxl not installed")
def test_excel_import_availability():
    """Test that openpyxl is available"""
    assert EXCEL_AVAILABLE is True
    import openpyxl
    assert openpyxl is not None


@pytest.mark.skipif(not EXCEL_AVAILABLE, reason="openpyxl not installed")
class TestSingletonPattern:
    """Test singleton pattern for Excel store"""

    @pytest.fixture(autouse=True)
    def setup_and_teardown(self, tmp_path, monkeypatch):
        """Setup and teardown for each test"""
        # Change to temp directory
        monkeypatch.chdir(tmp_path)
        # Clear singleton instances before each test
        ExcelStoreBase._instances.clear()
        yield
        # Cleanup after test
        ExcelStoreBase._instances.clear()

    def test_get_instance_returns_same_instance(self):
        """Test that get_instance returns the same instance for same parameters"""
        instance1 = ExcelStoreBase.get_instance("xhs", "search")
        instance2 = ExcelStoreBase.get_instance("xhs", "search")

        assert instance1 is instance2

    def test_get_instance_different_params_returns_different_instances(self):
        """Test that different parameters return different instances"""
        instance1 = ExcelStoreBase.get_instance("xhs", "search")
        instance2 = ExcelStoreBase.get_instance("xhs", "detail")
        instance3 = ExcelStoreBase.get_instance("douyin", "search")

        assert instance1 is not instance2
        assert instance1 is not instance3
        assert instance2 is not instance3

    @pytest.mark.asyncio
    async def test_singleton_preserves_data(self):
        """Test that singleton pattern preserves data across multiple calls"""
        # First call - store some content
        store1 = ExcelStoreBase.get_instance("test", "search")
        await store1.store_content({"note_id": "note1", "title": "Title 1"})

        # Second call - should get same instance with data
        store2 = ExcelStoreBase.get_instance("test", "search")
        await store2.store_content({"note_id": "note2", "title": "Title 2"})

        # Verify both items are in the same workbook
        assert store1 is store2
        assert store1.sheets["Contents"].max_row == 3

    def test_flush_all_saves_all_instances(self, tmp_path):
        """Test that flush_all saves all instances"""
        # Create multiple instances
        store1 = ExcelStoreBase.get_instance("platform1", "search")
        store2 = ExcelStoreBase.get_instance("platform2", "search")

        # Add data to each
        asyncio.run(store1.store_content({"note_id": "note1"}))
        asyncio.run(store2.store_content({"note_id": "note2"}))

        # Flush all
        ExcelStoreBase.flush_all()

        # Verify instances are cleared
        assert len(ExcelStoreBase._instances) == 0

        # Verify files were created
        assert store1.get_filename("content").exists()
        assert store2.get_filename("content").exists()

    def test_flush_all_clears_instances(self):
        """Test that flush_all clears the instances dictionary"""
        # Create an instance
        ExcelStoreBase.get_instance("test", "search")
        assert len(ExcelStoreBase._instances) == 1

        # Flush all
        ExcelStoreBase.flush_all()

        # Verify instances are cleared
        assert len(ExcelStoreBase._instances) == 0
