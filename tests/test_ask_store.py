# -*- coding: utf-8 -*-
# Copyright (c) 2025 relakkes@gmail.com
#
# This file is part of MediaCrawler project.
# Repository: https://github.com/NanmiCoder/MediaCrawler
# Licensed under NON-COMMERCIAL LEARNING LICENSE 1.1

import asyncio
import csv
import json

import config
import openpyxl
import pytest
from store.ask import (
    AskStoreFactory,
    _normalize_answer,
    _normalize_question,
    batch_update_ask_answers,
    update_ask_question,
)
from store.excel_store_base import ExcelStoreBase
from var import crawler_type_var


def question_data():
    return {
        "question_id": "104632506",
        "keyword": "小儿七星茶",
        "title": "小儿七星茶怎么喝",
        "description": "问题描述",
        "category": "小儿内科",
        "publish_time": "2022-06-14 02:23:31",
        "reply_count": 1,
        "question_url": "https://www.120ask.com/question/104632506.htm",
        "search_page": 3,
    }


def answer_data(comment_id: str = "1417651"):
    return {
        "comment_id": comment_id,
        "keyword": "小儿七星茶",
        "question_title": "小儿七星茶怎么喝",
        "question_description": "问题描述",
        "content": "病情分析：按说明书使用。",
        "doctor_name": "刘锋",
        "doctor_hospital": "九江市妇幼保健院",
        "doctor_title": "主治医师",
        "doctor_specialty": "小儿外科",
        "answer_time": "2022-06-14 06:43:15",
        "question_url": "https://www.120ask.com/question/104632506.htm",
    }


def test_ask_normalization_keeps_question_and_answer_fields():
    question = _normalize_question(question_data())
    answer = _normalize_answer("104632506", answer_data())

    assert question["question_id"] == "104632506"
    assert question["search_page"] == 3
    assert answer["question_id"] == "104632506"
    assert answer["question_title"] == "小儿七星茶怎么喝"
    assert answer["question_description"] == "问题描述"
    assert answer["doctor_hospital"] == "九江市妇幼保健院"
    assert "last_modify_ts" in question
    assert "last_modify_ts" in answer


def test_ask_store_rejects_database_output(monkeypatch):
    monkeypatch.setattr(config, "SAVE_DATA_OPTION", "db")

    with pytest.raises(ValueError, match="csv、json、jsonl 或 excel"):
        AskStoreFactory.create_store()


@pytest.mark.parametrize("save_option", ["json", "jsonl", "csv"])
def test_ask_file_formats_persist_question_and_each_answer(
    monkeypatch,
    tmp_path,
    save_option,
):
    monkeypatch.setattr(config, "SAVE_DATA_OPTION", save_option)
    monkeypatch.setattr(config, "SAVE_DATA_PATH", str(tmp_path))
    crawler_type_var.set("search")

    asyncio.run(update_ask_question(question_data()))
    asyncio.run(
        batch_update_ask_answers(
            "104632506",
            [answer_data("1"), answer_data("2")],
        )
    )

    content_file = next((tmp_path / "ask" / save_option).glob("ask_search_content_*"))
    comment_file = next((tmp_path / "ask" / save_option).glob("ask_search_comment_*"))
    if save_option == "json":
        content_records = json.loads(content_file.read_text(encoding="utf-8"))
        comment_records = json.loads(comment_file.read_text(encoding="utf-8"))
    elif save_option == "jsonl":
        content_records = [
            json.loads(line)
            for line in content_file.read_text(encoding="utf-8").splitlines()
        ]
        comment_records = [
            json.loads(line)
            for line in comment_file.read_text(encoding="utf-8").splitlines()
        ]
    else:
        with content_file.open(encoding="utf-8-sig", newline="") as file:
            content_records = list(csv.DictReader(file))
        with comment_file.open(encoding="utf-8-sig", newline="") as file:
            comment_records = list(csv.DictReader(file))

    assert len(content_records) == 1
    assert len(comment_records) == 2
    assert content_records[0]["question_id"] == "104632506"
    assert [record["comment_id"] for record in comment_records] == ["1", "2"]
    assert all(
        record["question_title"] == "小儿七星茶怎么喝"
        for record in comment_records
    )
    assert all(
        record["question_description"] == "问题描述"
        for record in comment_records
    )


def test_ask_excel_persists_question_and_answer_immediately(monkeypatch, tmp_path):
    monkeypatch.setattr(config, "SAVE_DATA_OPTION", "excel")
    monkeypatch.setattr(config, "SAVE_DATA_PATH", str(tmp_path))
    crawler_type_var.set("search")
    ExcelStoreBase._instances.clear()

    asyncio.run(update_ask_question(question_data()))
    asyncio.run(batch_update_ask_answers("104632506", [answer_data()]))

    content_file = next((tmp_path / "ask").glob("ask_search_content_*.xlsx"))
    comment_file = next((tmp_path / "ask").glob("ask_search_comment_*.xlsx"))
    content_book = openpyxl.load_workbook(content_file)
    comment_book = openpyxl.load_workbook(comment_file)

    assert content_book["Contents"].max_row == 2
    assert comment_book["Comments"].max_row == 2
    content_book.close()
    comment_book.close()
    ExcelStoreBase._instances.clear()
