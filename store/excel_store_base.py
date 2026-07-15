# -*- coding: utf-8 -*-
# Copyright (c) 2025 relakkes@gmail.com
#
# This file is part of MediaCrawler project.
# Repository: https://github.com/NanmiCoder/MediaCrawler/blob/main/store/excel_store_base.py
# GitHub: https://github.com/NanmiCoder
# Licensed under NON-COMMERCIAL LEARNING LICENSE 1.1
#
# 声明：本代码仅供学习和研究目的使用。使用者应遵守以下原则：
# 1. 不得用于任何商业用途。
# 2. 使用时应遵守目标平台的使用条款和robots.txt规则。
# 3. 不得进行大规模爬取或对平台造成运营干扰。
# 4. 应合理控制请求频率，避免给目标平台带来不必要的负担。
# 5. 不得用于任何非法或不当的用途。
#
# 详细许可条款请参阅项目根目录下的LICENSE文件。
# 使用本代码即表示您同意遵守上述原则和LICENSE中的所有条款。

# 声明:本代码仅供学习和研究目的使用。使用者应遵守以下原则:
# 1. 不得用于任何商业用途。
# 2. 使用时应遵守目标平台的使用条款和robots.txt规则。
# 3. 不得进行大规模爬取或对平台造成运营干扰。
# 4. 应合理控制请求频率,避免给目标平台带来不必要的负担。
# 5. 不得用于任何非法或不当的用途。
#
# 详细许可条款请参阅项目根目录下的LICENSE文件。
# 使用本代码即表示您同意遵守上述原则和LICENSE中的所有条款。

"""
Excel Store Base Implementation
Provides Excel export functionality for crawled data with formatted sheets
"""

import threading
from typing import Dict, List, Any
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from base.base_crawler import AbstractStore
from tools import utils
from tools.file_naming import build_data_filename, normalize_data_type
import config


class ExcelStoreBase(AbstractStore):
    """
    Base class for Excel storage implementation
    Provides formatted Excel export with multiple sheets for contents, comments, and creators
    Uses singleton pattern to maintain state across multiple store calls
    """

    # Class-level singleton management
    _instances: Dict[str, "ExcelStoreBase"] = {}
    _lock = threading.Lock()

    @classmethod
    def get_instance(cls, platform: str, crawler_type: str) -> "ExcelStoreBase":
        """
        Get or create a singleton instance for the given platform and crawler type

        Args:
            platform: Platform name (xhs, dy, ks, etc.)
            crawler_type: Type of crawler (search, detail, creator)

        Returns:
            ExcelStoreBase instance
        """
        key = f"{platform}_{crawler_type}"
        with cls._lock:
            if key not in cls._instances:
                cls._instances[key] = cls(platform, crawler_type)
            return cls._instances[key]

    @classmethod
    def flush_all(cls):
        """
        Flush all Excel store instances and save to files
        Should be called at the end of crawler execution
        """
        with cls._lock:
            for key, instance in cls._instances.items():
                try:
                    instance.flush()
                    utils.logger.info(f"[ExcelStoreBase] Flushed instance: {key}")
                except Exception as e:
                    utils.logger.error(f"[ExcelStoreBase] Error flushing {key}: {e}")
            cls._instances.clear()

    def __init__(self, platform: str, crawler_type: str = "search"):
        """
        Initialize Excel store

        Args:
            platform: Platform name (xhs, dy, ks, etc.)
            crawler_type: Type of crawler (search, detail, creator)
        """
        if not EXCEL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

        super().__init__()
        self.platform = platform
        self.crawler_type = crawler_type

        # Create data directory
        if config.SAVE_DATA_PATH:
            self.data_dir = Path(config.SAVE_DATA_PATH) / platform
        else:
            self.data_dir = Path("data") / platform
        self.data_dir.mkdir(parents=True, exist_ok=True)

        self.workbooks: Dict[str, Any] = {}
        self.sheets: Dict[str, Any] = {}
        self.headers_written: set[str] = set()
        self.filenames = {
            data_type: self.data_dir
            / build_data_filename(platform, crawler_type, data_type, "xlsx")
            for data_type in ("content", "comment")
        }

        utils.logger.info(
            f"[ExcelStoreBase] Initialized Excel export: {self.filenames}"
        )

    def get_filename(self, item_type: str) -> Path:
        """Return the content or comment workbook path for this run."""
        return self.filenames[normalize_data_type(item_type)]

    def _get_sheet(self, item_type: str, sheet_name: str):
        """Create a workbook and sheet lazily for one normalized data type."""
        data_type = normalize_data_type(item_type)
        workbook = self.workbooks.get(data_type)
        if workbook is None:
            workbook = openpyxl.Workbook()
            workbook.active.title = sheet_name
            self.workbooks[data_type] = workbook
            sheet = workbook.active
        elif sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)
        self.sheets[sheet_name] = sheet
        return sheet

    def _apply_header_style(self, sheet, row_num: int = 1):
        """
        Apply formatting to header row

        Args:
            sheet: Worksheet object
            row_num: Row number for headers (default: 1)
        """
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in sheet[row_num]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

    def _auto_adjust_column_width(self, sheet):
        """
        Auto-adjust column widths based on content

        Args:
            sheet: Worksheet object
        """
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (TypeError, AttributeError):
                    pass

            # Set width with min/max constraints
            adjusted_width = min(max(max_length + 2, 10), 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _write_headers(self, sheet, headers: List[str]):
        """
        Write headers to sheet

        Args:
            sheet: Worksheet object
            headers: List of header names
        """
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        self._apply_header_style(sheet)

    def _write_row(self, sheet, data: Dict[str, Any], headers: List[str]):
        """
        Write data row to sheet

        Args:
            sheet: Worksheet object
            data: Data dictionary
            headers: List of header names (defines column order)
        """
        row_num = sheet.max_row + 1

        for col_num, header in enumerate(headers, 1):
            value = data.get(header, "")

            # Handle different data types
            if isinstance(value, (list, dict)):
                value = str(value)
            elif value is None:
                value = ""

            cell = sheet.cell(row=row_num, column=col_num, value=value)

            # Apply basic formatting
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    async def store_content(self, content_item: Dict):
        """
        Store content data to Excel

        Args:
            content_item: Content data dictionary
        """
        sheet = self._get_sheet("content", "Contents")
        headers = list(content_item.keys())

        if "Contents" not in self.headers_written:
            self._write_headers(sheet, headers)
            self.headers_written.add("Contents")

        self._write_row(sheet, content_item, headers)

        # Get ID from various possible field names
        content_id = content_item.get('note_id') or content_item.get('aweme_id') or content_item.get('video_id') or content_item.get('content_id') or 'N/A'
        utils.logger.info(f"[ExcelStoreBase] Stored content to Excel: {content_id}")

    async def store_comment(self, comment_item: Dict):
        """
        Store comment data to Excel

        Args:
            comment_item: Comment data dictionary
        """
        sheet = self._get_sheet("comment", "Comments")
        headers = list(comment_item.keys())

        if "Comments" not in self.headers_written:
            self._write_headers(sheet, headers)
            self.headers_written.add("Comments")

        self._write_row(sheet, comment_item, headers)

        # Persist every completed comment immediately so a later request failure
        # cannot discard comments already returned by the platform.
        self.workbooks["comment"].save(self.filenames["comment"])

        utils.logger.info(f"[ExcelStoreBase] Stored comment to Excel: {comment_item.get('comment_id', 'N/A')}")

    async def store_creator(self, creator: Dict):
        """
        Store creator data to Excel

        Args:
            creator: Creator data dictionary
        """
        sheet = self._get_sheet("content", "Creators")
        headers = list(creator.keys())

        if "Creators" not in self.headers_written:
            self._write_headers(sheet, headers)
            self.headers_written.add("Creators")

        self._write_row(sheet, creator, headers)

        utils.logger.info(f"[ExcelStoreBase] Stored creator to Excel: {creator.get('user_id', 'N/A')}")

    async def store_contact(self, contact_item: Dict):
        """
        Store contact data to Excel (for platforms like Bilibili)

        Args:
            contact_item: Contact data dictionary
        """
        sheet = self._get_sheet("content", "Contacts")
        headers = list(contact_item.keys())

        if "Contacts" not in self.headers_written:
            self._write_headers(sheet, headers)
            self.headers_written.add("Contacts")

        self._write_row(sheet, contact_item, headers)

        utils.logger.info(f"[ExcelStoreBase] Stored contact to Excel: up_id={contact_item.get('up_id', 'N/A')}, fan_id={contact_item.get('fan_id', 'N/A')}")

    async def store_dynamic(self, dynamic_item: Dict):
        """
        Store dynamic data to Excel (for platforms like Bilibili)

        Args:
            dynamic_item: Dynamic data dictionary
        """
        sheet = self._get_sheet("content", "Dynamics")
        headers = list(dynamic_item.keys())

        if "Dynamics" not in self.headers_written:
            self._write_headers(sheet, headers)
            self.headers_written.add("Dynamics")

        self._write_row(sheet, dynamic_item, headers)

        utils.logger.info(f"[ExcelStoreBase] Stored dynamic to Excel: {dynamic_item.get('dynamic_id', 'N/A')}")

    def flush(self):
        """
        Save workbook to file
        """
        try:
            if not self.workbooks:
                utils.logger.info(
                    "[ExcelStoreBase] No data to save, skipping file creation"
                )
                return

            for data_type, workbook in self.workbooks.items():
                for sheet in workbook.worksheets:
                    self._auto_adjust_column_width(sheet)
                filename = self.filenames[data_type]
                workbook.save(filename)
                utils.logger.info(
                    f"[ExcelStoreBase] Excel file saved successfully: {filename}"
                )

        except Exception as e:
            utils.logger.error(f"[ExcelStoreBase] Error saving Excel file: {e}")
            raise
